import os
import sqlite3
from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'data', 'agrishow.db')
XLSX_PATH = os.path.join(BASE_DIR, 'data', 'Agrishow_Machine_Control_Sheet.xlsm')

DEALERS = [
    'CEQUIP', 'DAMAQ', 'JUMASA', 'JUMASA NORTE', 'MEVOS', 'MPM',
    'NORDESTE', 'PRIORI', 'RR', 'SARANDI', 'SERPEMA', 'TRACSUL', 'TRACTORBEL'
]

# ---------------- SCHEMA ----------------
def criar_schema(conn):
    c = conn.cursor()

    c.executescript("""
    CREATE TABLE IF NOT EXISTS maquinas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        modelo TEXT NOT NULL,
        sap TEXT UNIQUE NOT NULL,
        estoque_inicial_15 INTEGER DEFAULT 0,
        estoque_inicial_30 INTEGER DEFAULT 0,
        estoque_inicial_60 INTEGER DEFAULT 0
    );

    CREATE TABLE IF NOT EXISTS dealers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nome TEXT UNIQUE NOT NULL
    );

    CREATE TABLE IF NOT EXISTS pedidos (
        id TEXT PRIMARY KEY,
        data_hora TEXT NOT NULL,
        dealer TEXT NOT NULL,
        funcionario TEXT NOT NULL,
        modelo TEXT NOT NULL,
        sap TEXT NOT NULL,
        quantidade INTEGER NOT NULL,
        prazo INTEGER NOT NULL,
        anexo_filename TEXT NOT NULL,
        status TEXT DEFAULT 'ACEITO'
    );
    """)

    conn.commit()

# ---------------- DEALERS ----------------
def popular_dealers(conn):
    c = conn.cursor()
    for d in DEALERS:
        c.execute("INSERT OR IGNORE INTO dealers (nome) VALUES (?)", (d,))
    conn.commit()

# ---------------- UPSERT MANUAL ----------------
def upsert_maquina(conn, modelo, sap, e15, e30, e60):
    c = conn.cursor()

    # atualiza se existir
    c.execute("""
        UPDATE maquinas
        SET modelo=?, estoque_inicial_15=?, estoque_inicial_30=?, estoque_inicial_60=?
        WHERE sap=?
    """, (modelo, int(e15), int(e30), int(e60), sap))

    # se não existia, insere
    if c.rowcount == 0:
        c.execute("""
            INSERT INTO maquinas (modelo, sap, estoque_inicial_15, estoque_inicial_30, estoque_inicial_60)
            VALUES (?, ?, ?, ?, ?)
        """, (modelo, sap, int(e15), int(e30), int(e60)))

# ---------------- MAQUINAS ----------------
def popular_maquinas(conn):
    c = conn.cursor()

    # -------- FALLBACK (SE EXCEL FALHAR) --------
    exemplos = [
        ('4160D', '35F01190003B001', 0, 8, 57),
        ('6612E', '23F00700020B003', 15, 0, 0),
        ('6612E', '23F00700022B002', 39, 0, 0),
        ('818H', '60F01100006B002', 0, 0, 10),
        ('835T', '62F02200001B001', 8, 0, 117),
        ('835T', '62F02200002B001', 11, 0, 0),
        ('838T', '62F02310001B001', 5, 0, 0),
        ('848T', '64F08560014B001', 6, 0, 0),
        ('908E', '06F0187C032B001', 10, 18, 18),
        ('908E', '06F0187C049B001', 20, 0, 0),
        ('913E', '08F00550010B002', 0, 5, 16),
        ('913E', '08F00550011B002', 0, 25, 0),
        ('915E', '08F00630020B001', 1, 0, 42),
        ('915E', '08F00630021B001', 3, 2, 20),
        ('915E', '08F00630022B001', 1, 2, 0),
        ('915E', '08F00630023B001', 4, 1, 10),
        ('922E', '10F0084C075B001', 0, 6, 26),
        ('922E', '10F0084C076B001', 0, 6, 6),
        ('922E', '10F0084C077B001', 0, 0, 35),
    ]

    # tenta Excel
    if os.path.exists(XLSX_PATH):
        try:
            wb = load_workbook(XLSX_PATH, data_only=True)
            ws = wb['Inicial']

            for row in ws.iter_rows(min_row=2, values_only=True):
                modelo, sap = row[0], row[1]
                if not modelo or not sap:
                    continue

                e15 = row[2] or 0
                e30 = row[3] or 0
                e60 = row[4] or 0

                upsert_maquina(conn, modelo, sap, e15, e30, e60)

            conn.commit()
            print("Dados carregados do Excel")
            return

        except Exception as e:
            print(f"Erro no Excel, usando fallback: {e}")

    # fallback
    for m in exemplos:
        upsert_maquina(conn, *m)

    conn.commit()
    print("Dados carregados do fallback")

# ---------------- MAIN ----------------
def main():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)

    conn = sqlite3.connect(DB_PATH)

    criar_schema(conn)
    popular_dealers(conn)
    popular_maquinas(conn)

    total_m = conn.execute("SELECT COUNT(*) FROM maquinas").fetchone()[0]
    total_d = conn.execute("SELECT COUNT(*) FROM dealers").fetchone()[0]

    print(f"Banco atualizado: {total_m} maquinas | {total_d} dealers")

    conn.close()

if __name__ == '__main__':
    main()
